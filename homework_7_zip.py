from zipfile import ZipFile
from pypdf import PdfReader
from openpyxl import load_workbook
from io import BytesIO

def test_arc_pdf(create_archive_file):
    with ZipFile("files/archive.zip", "r") as archive:
        print(archive.namelist())
        with archive.open('pdf_list.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            first_page_text = reader.pages[0].extract_text()
        assert "понимания информации, а т акж е для проверки знаний по определенной т еме" in \
               reader.pages[0].extract_text()


def test_arc_csv(create_archive_file):
    with ZipFile("files/archive.zip", "r") as archive:
        print(archive.namelist())
        with archive.open('user_list.csv') as user_list:
            content = user_list.read().decode('utf-8')
            print(content)
            assert "ksergeev@company.ru" in content


def test_arc_xlsx(create_archive_file):
    with ZipFile("files/archive.zip", "r") as archive:
        print(archive.namelist())
        with archive.open('user_list_for_xlsx.xlsx') as user_list_for_xlsx:
            wb = load_workbook(filename=BytesIO(user_list_for_xlsx.read()))
            ws = wb.active
            found = False
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and "ksergeev@company.ru" in str(cell.value):
                        found = True
                        break
                if found:
                    break
            assert found, "Email ksergeev@company.ru не найден в файле"

